import streamlit as st
import pandas as pd
import numpy as np
from sklearn.ensemble import RandomForestRegressor, RandomForestClassifier
from sklearn.preprocessing import MinMaxScaler
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_absolute_error, r2_score, accuracy_score
import warnings
warnings.filterwarnings('ignore')
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from io import BytesIO
import base64

# ==================== CONFIGURACIÓN INICIAL ====================
st.set_page_config(
    page_title="Inseguridad Alimentaria Lima 2024-2035",
    page_icon="🍽️",
    layout="wide",
    initial_sidebar_state="expanded"
)

# ==================== CSS PERSONALIZADO ====================
custom_css = """
<style>
    * {
        margin: 0;
        padding: 0;
        box-sizing: border-box;
    }
    
    body {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif;
    }
    
    .main {
        background: white;
        border-radius: 15px;
    }
    
    /* Títulos */
    h1, h2, h3 {
        color: #1a472a;
        font-weight: 700;
        margin-bottom: 20px;
    }
    
    h1 {
        font-size: 2.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.1);
    }
    
    h2 {
        font-size: 1.8rem;
        border-bottom: 3px solid #2d8659;
        padding-bottom: 10px;
    }
    
    /* Cards KPI */
    .kpi-card {
        background: linear-gradient(135deg, #2d8659 0%, #4caf50 100%);
        color: white;
        padding: 25px;
        border-radius: 12px;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1);
        text-align: center;
        transition: all 0.3s ease;
        border-left: 5px solid #ff9800;
    }
    
    .kpi-card:hover {
        transform: translateY(-5px);
        box-shadow: 0 8px 25px rgba(0,0,0,0.15);
    }
    
    .kpi-value {
        font-size: 2.5rem;
        font-weight: 700;
        margin: 10px 0;
    }
    
    .kpi-label {
        font-size: 0.95rem;
        opacity: 0.9;
        text-transform: uppercase;
        letter-spacing: 1px;
    }
    
    /* Alerts */
    .alert-danger {
        background: #ffebee;
        border-left: 5px solid #f44336;
        color: #c62828;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
    
    .alert-warning {
        background: #fff3e0;
        border-left: 5px solid #ff9800;
        color: #e65100;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
    
    .alert-success {
        background: #e8f5e9;
        border-left: 5px solid #4caf50;
        color: #1b5e20;
        padding: 15px;
        border-radius: 5px;
        margin: 10px 0;
    }
    
    /* Tablas */
    .dataframe {
        border-collapse: collapse;
        width: 100%;
    }
    
    .dataframe thead {
        background: #1a472a;
        color: white;
    }
    
    .dataframe tbody tr:nth-child(odd) {
        background: #f5f5f5;
    }
    
    .dataframe tbody tr:hover {
        background: #e8f5e9;
    }
    
    /* Botones */
    button {
        background: linear-gradient(135deg, #2d8659 0%, #4caf50 100%);
        color: white;
        border: none;
        padding: 12px 24px;
        border-radius: 8px;
        font-weight: 600;
        cursor: pointer;
        transition: all 0.3s ease;
    }
    
    button:hover {
        box-shadow: 0 4px 12px rgba(45, 134, 89, 0.4);
        transform: translateY(-2px);
    }
    
    /* Inputs */
    input, select {
        border: 2px solid #e0e0e0;
        border-radius: 8px;
        padding: 10px;
        font-size: 0.95rem;
    }
    
    input:focus, select:focus {
        border-color: #2d8659;
        box-shadow: 0 0 8px rgba(45, 134, 89, 0.2);
    }
    
    /* Barras de progreso */
    .progress-bar {
        height: 25px;
        border-radius: 10px;
        background: #e0e0e0;
        overflow: hidden;
        margin: 10px 0;
    }
    
    .progress-fill {
        height: 100%;
        display: flex;
        align-items: center;
        justify-content: center;
        color: white;
        font-weight: 600;
        font-size: 0.85rem;
    }
    
    .progress-fill.muy-alto {
        background: linear-gradient(90deg, #d32f2f, #f44336);
    }
    
    .progress-fill.alto {
        background: linear-gradient(90deg, #f57c00, #ff9800);
    }
    
    .progress-fill.medio {
        background: linear-gradient(90deg, #fbc02d, #fdd835);
    }
    
    .progress-fill.bajo {
        background: linear-gradient(90deg, #388e3c, #4caf50);
    }
    
    /* Sidebar */
    [data-testid="stSidebar"] {
        background: linear-gradient(180deg, #1a472a 0%, #2d8659 100%);
    }
    
    [data-testid="stSidebar"] h1, 
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        color: white;
    }
    
    [data-testid="stSidebar"] label {
        color: white;
    }
</style>
"""

st.markdown(custom_css, unsafe_allow_html=True)

# ==================== FUNCIONES AUXILIARES ====================

@st.cache_data
def cargar_datos():
    """Carga y prepara los datos"""
    try:
        df = pd.read_excel('Entregable_3_Dataset_Transformado_Inseguridad_Alimentaria.xlsx')
        return df
    except Exception as e:
        st.error(f"Error al cargar datos: {e}")
        return None

def normalizar_datos(df):
    """Normaliza variables críticas"""
    scaler = MinMaxScaler()
    columnas_a_normalizar = ['Ingresos_Laborales', 'Gasto_Alimentario', 
                             'Indice_Vulnerabilidad', 'Indice_Privacion_Multidimensional']
    
    df_normalizado = df.copy()
    for col in columnas_a_normalizar:
        if col in df.columns:
            df_normalizado[col] = scaler.fit_transform(df[[col]])
    
    return df_normalizado

def calcular_iria(ingresos, gasto_alimentario, vulnerabilidad, privacion):
    """
    Calcula el Índice de Riesgo de Inseguridad Alimentaria (IRIA)
    
    Ponderación:
    - Ingresos: 30% (inverso: menos ingresos = más riesgo)
    - Gasto Alimentario: 25%
    - Vulnerabilidad: 25%
    - Privación: 20%
    """
    iria = (
        (1 - ingresos) * 0.30 +  # Inverso porque menos ingresos = más riesgo
        (1 - gasto_alimentario) * 0.25 +
        vulnerabilidad * 0.25 +
        privacion * 0.20
    )
    return iria

def nivel_iria_a_texto(iria):
    """Convierte valor IRIA a nivel de riesgo"""
    if iria >= 0.75:
        return "Muy Alto"
    elif iria >= 0.50:
        return "Alto"
    elif iria >= 0.25:
        return "Medio"
    else:
        return "Bajo"

def nivel_iria_a_color(nivel):
    """Retorna color según nivel de IRIA"""
    colores = {
        "Muy Alto": "#d32f2f",
        "Alto": "#ff9800",
        "Medio": "#fdd835",
        "Bajo": "#4caf50"
    }
    return colores.get(nivel, "#999")

def entrenar_modelos(df_norm):
    """Entrena los modelos de ML"""
    # Preparar datos
    caracteristicas = ['Ingresos_Laborales', 'Gasto_Alimentario', 
                      'Indice_Vulnerabilidad', 'Indice_Privacion_Multidimensional']
    
    X = df_norm[caracteristicas].fillna(0)
    y_reg = df_norm['Probabilidad_Inseguridad'].fillna(0)
    y_clf = df_norm['Clasificacion_Riesgo'].fillna('Bajo')
    
    # Codificar clasificación
    clasificaciones = {'Bajo': 0, 'Medio': 1, 'Alto': 2, 'Muy Alto': 3}
    y_clf_encoded = y_clf.map(clasificaciones).fillna(0)
    
    # División train-test
    X_train, X_test, y_reg_train, y_reg_test = train_test_split(
        X, y_reg, test_size=0.2, random_state=42
    )
    _, _, y_clf_train, y_clf_test = train_test_split(
        X, y_clf_encoded, test_size=0.2, random_state=42
    )
    
    # Entrenar Regresor
    reg_model = RandomForestRegressor(n_estimators=100, random_state=42, max_depth=10)
    reg_model.fit(X_train, y_reg_train)
    
    # Entrenar Clasificador
    clf_model = RandomForestClassifier(n_estimators=100, random_state=42, max_depth=10)
    clf_model.fit(X_train, y_clf_train)
    
    # Métricas
    y_reg_pred = reg_model.predict(X_test)
    y_clf_pred = clf_model.predict(X_test)
    
    mae = mean_absolute_error(y_reg_test, y_reg_pred)
    r2 = r2_score(y_reg_test, y_reg_pred)
    accuracy = accuracy_score(y_clf_test, y_clf_pred)
    
    return {
        'regressor': reg_model,
        'classifier': clf_model,
        'mae': mae,
        'r2': r2,
        'accuracy': accuracy,
        'caracteristicas': caracteristicas
    }

def proyectar_riesgo(df, modelos, distrito, año_inicio, año_fin):
    """Proyecta riesgo para años futuros"""
    proyecciones = []
    
    # Datos del distrito
    distrito_data = df[df['Distrito'] == distrito].iloc[0] if distrito in df['Distrito'].values else None
    
    if distrito_data is None:
        return pd.DataFrame()
    
    for año in range(año_inicio, año_fin + 1):
        años_transcurridos = año - 2024
        presion_economica = 1 - (0.02 * años_transcurridos)  # Deterioro del 2% anual
        
        # Ajustar características
        ingresos_ajustados = distrito_data['Ingresos_Laborales'] * presion_economica
        gasto_ajustado = distrito_data['Gasto_Alimentario'] * (1 - 0.01 * años_transcurridos)
        vulnerabilidad_ajustada = min(1.0, distrito_data['Indice_Vulnerabilidad'] + 0.01 * años_transcurridos)
        privacion_ajustada = min(1.0, distrito_data['Indice_Privacion_Multidimensional'] + 0.015 * años_transcurridos)
        
        # Calcular IRIA
        iria = calcular_iria(ingresos_ajustados, gasto_ajustado, vulnerabilidad_ajustada, privacion_ajustada)
        nivel_iria = nivel_iria_a_texto(iria)
        
        # Predecir con modelos
        X_pred = np.array([[ingresos_ajustados, gasto_ajustado, vulnerabilidad_ajustada, privacion_ajustada]])
        probabilidad = modelos['regressor'].predict(X_pred)[0]
        probabilidad = np.clip(probabilidad, 0, 1)
        
        proyecciones.append({
            'Año': año,
            'Distrito': distrito,
            'IRIA': iria,
            'Nivel_Riesgo': nivel_iria,
            'Probabilidad': probabilidad,
            'Ingresos': ingresos_ajustados,
            'Gasto_Alimentario': gasto_ajustado,
            'Vulnerabilidad': vulnerabilidad_ajustada,
            'Privacion': privacion_ajustada
        })
    
    return pd.DataFrame(proyecciones)

def crear_ranking_distritos(df, modelos, año):
    """Crea ranking de distritos por IRIA"""
    ranking = []
    
    for distrito in df['Distrito'].unique():
        distrito_data = df[df['Distrito'] == distrito].iloc[0]
        
        # Calcular IRIA actual
        iria = calcular_iria(
            distrito_data['Ingresos_Laborales'],
            distrito_data['Gasto_Alimentario'],
            distrito_data['Indice_Vulnerabilidad'],
            distrito_data['Indice_Privacion_Multidimensional']
        )
        
        nivel_riesgo = nivel_iria_a_texto(iria)
        
        # Predicción
        X_pred = np.array([[
            distrito_data['Ingresos_Laborales'],
            distrito_data['Gasto_Alimentario'],
            distrito_data['Indice_Vulnerabilidad'],
            distrito_data['Indice_Privacion_Multidimensional']
        ]])
        probabilidad = modelos['regressor'].predict(X_pred)[0]
        probabilidad = np.clip(probabilidad, 0, 1)
        
        ranking.append({
            'Distrito': distrito,
            'IRIA': round(iria, 4),
            'Nivel_Riesgo': nivel_riesgo,
            'Probabilidad': round(probabilidad, 4),
            'Población': int(distrito_data.get('Poblacion', 0))
        })
    
    ranking_df = pd.DataFrame(ranking).sort_values('IRIA', ascending=False)
    return ranking_df.reset_index(drop=True)

def crear_grafica_distribucion_iria(df_ranking):
    """Crea gráfica didáctica de distribución de IRIA"""
    fig, axes = plt.subplots(2, 2, figsize=(14, 10))
    fig.suptitle('Análisis Didáctico de Inseguridad Alimentaria por Distrito', 
                 fontsize=16, fontweight='bold', color='#1a472a')
    
    # 1. Gráfica de barras horizontales (IRIA por distrito)
    ax1 = axes[0, 0]
    top_10 = df_ranking.head(10).sort_values('IRIA', ascending=True)
    colores = [nivel_iria_a_color(nivel) for nivel in top_10['Nivel_Riesgo']]
    ax1.barh(range(len(top_10)), top_10['IRIA'], color=colores, edgecolor='black', linewidth=1.5)
    ax1.set_yticks(range(len(top_10)))
    ax1.set_yticklabels(top_10['Distrito'], fontsize=9)
    ax1.set_xlabel('Índice IRIA', fontweight='bold')
    ax1.set_title('Top 10 Distritos con Mayor Riesgo', fontweight='bold', color='#1a472a')
    ax1.grid(axis='x', alpha=0.3, linestyle='--')
    ax1.set_xlim(0, 1)
    
    # 2. Gráfica de pastel (Distribución por nivel)
    ax2 = axes[0, 1]
    conteo_niveles = df_ranking['Nivel_Riesgo'].value_counts()
    orden_niveles = ['Muy Alto', 'Alto', 'Medio', 'Bajo']
    conteo_niveles = conteo_niveles.reindex([n for n in orden_niveles if n in conteo_niveles.index])
    
    colores_pie = [nivel_iria_a_color(nivel) for nivel in conteo_niveles.index]
    wedges, texts, autotexts = ax2.pie(
        conteo_niveles.values,
        labels=conteo_niveles.index,
        colors=colores_pie,
        autopct='%1.1f%%',
        startangle=90,
        textprops={'fontsize': 10, 'fontweight': 'bold'}
    )
    ax2.set_title('Distribución de Distritos por Nivel de Riesgo', fontweight='bold', color='#1a472a')
    
    # 3. Gráfica de dispersión (IRIA vs Probabilidad)
    ax3 = axes[1, 0]
    scatter = ax3.scatter(df_ranking['IRIA'], df_ranking['Probabilidad'], 
                         s=df_ranking['Población']/100, 
                         c=[nivel_iria_a_color(n) for n in df_ranking['Nivel_Riesgo']],
                         alpha=0.6, edgecolors='black', linewidth=1.5)
    ax3.set_xlabel('Índice IRIA', fontweight='bold')
    ax3.set_ylabel('Probabilidad de Inseguridad', fontweight='bold')
    ax3.set_title('Relación IRIA vs Probabilidad\n(Tamaño = Población)', fontweight='bold', color='#1a472a')
    ax3.grid(True, alpha=0.3, linestyle='--')
    ax3.set_xlim(0, 1)
    ax3.set_ylim(0, 1)
    
    # 4. Gráfica de barras apiladas (Conteo de distritos)
    ax4 = axes[1, 1]
    nivel_counts = pd.DataFrame({
        'Nivel': orden_niveles,
        'Cantidad': [len(df_ranking[df_ranking['Nivel_Riesgo'] == nivel]) for nivel in orden_niveles]
    })
    nivel_counts = nivel_counts[nivel_counts['Cantidad'] > 0]
    
    barras = ax4.bar(nivel_counts['Nivel'], nivel_counts['Cantidad'],
                     color=[nivel_iria_a_color(n) for n in nivel_counts['Nivel']],
                     edgecolor='black', linewidth=1.5)
    ax4.set_ylabel('Cantidad de Distritos', fontweight='bold')
    ax4.set_title('Cantidad de Distritos por Nivel de Riesgo', fontweight='bold', color='#1a472a')
    ax4.grid(axis='y', alpha=0.3, linestyle='--')
    
    # Añadir valores en las barras
    for barra in barras:
        altura = barra.get_height()
        ax4.text(barra.get_x() + barra.get_width()/2., altura,
                f'{int(altura)}',
                ha='center', va='bottom', fontweight='bold', fontsize=10)
    
    plt.tight_layout()
    return fig

def exportar_a_excel(df_ranking, modelos_info):
    """Exporta ranking a Excel"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Ranking Distritos"
    
    # Encabezados
    headers = ['Ranking', 'Distrito', 'IRIA', 'Nivel de Riesgo', 'Probabilidad', 'Población']
    header_fill = PatternFill(start_color="1a472a", end_color="1a472a", fill_type="solid")
    header_font = Font(bold=True, color="ffffff", size=11)
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # Datos
    for idx, row in df_ranking.iterrows():
        ws.cell(row=idx+2, column=1, value=idx+1)
        ws.cell(row=idx+2, column=2, value=row['Distrito'])
        ws.cell(row=idx+2, column=3, value=round(row['IRIA'], 4))
        ws.cell(row=idx+2, column=4, value=row['Nivel_Riesgo'])
        ws.cell(row=idx+2, column=5, value=round(row['Probabilidad'], 4))
        ws.cell(row=idx+2, column=6, value=row['Población'])
        
        # Colorear según nivel de riesgo
        color_hex = nivel_iria_a_color(row['Nivel_Riesgo']).lstrip('#')
        fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
        for col in range(1, 7):
            ws.cell(row=idx+2, column=col).fill = fill
    
    # Ajustar ancho de columnas
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    
    # Información de modelos
    ws2 = wb.create_sheet("Información Modelos")
    ws2['A1'] = "Métrica"
    ws2['B1'] = "Valor"
    
    metricas = [
        ["MAE", f"{modelos_info['mae']:.4f}"],
        ["R² Score", f"{modelos_info['r2']:.4f}"],
        ["Accuracy", f"{modelos_info['accuracy']:.4f}"],
        ["Fecha de Generación", datetime.now().strftime("%Y-%m-%d %H:%M:%S")]
    ]
    
    for idx, (metrica, valor) in enumerate(metricas, 2):
        ws2[f'A{idx}'] = metrica
        ws2[f'B{idx}'] = valor
    
    # Guardar a bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

# ==================== INTERFAZ PRINCIPAL ====================

# Header
col1, col2 = st.columns([0.7, 0.3])

with col1:
    st.markdown("""
    # Inseguridad Alimentaria en Lima Metropolitana
    ### Predicción y Clasificación de Riesgo (2024-2035)
    """)

with col2:
    st.image("https://www.lima.gob.pe/images/logo-munilima.png", width=150)

st.markdown("---")

# Cargar datos
df = cargar_datos()

if df is not None:
    df_normalizado = normalizar_datos(df)
    modelos = entrenar_modelos(df_normalizado)
    
    # ==================== SIDEBAR ====================
    with st.sidebar:
        st.markdown("## Configuración")
        
        tab_tipo = st.radio(
            "Selecciona el tipo de análisis:",
            [" Ranking de Distritos", " Proyecciones Futuras", " Análisis Detallado"]
        )
        
        año_analisis = st.slider(
            "Año de análisis",
            min_value=2024,
            max_value=2035,
            value=2024,
            step=1
        )
        
        st.markdown("---")
        st.markdown("### Información del Modelo")
        st.info(f"""
        **Métricas de Rendimiento:**
        - MAE: {modelos['mae']:.4f}
        - R² Score: {modelos['r2']:.4f}
        - Accuracy: {modelos['accuracy']:.4f}
        """)
    
    # ==================== PESTAÑA 1: RANKING DE DISTRITOS ====================
    if tab_tipo == " Ranking de Distritos":
        st.markdown("## Ranking de Distritos - Inseguridad Alimentaria")
        
        ranking_df = crear_ranking_distritos(df_normalizado, modelos, año_analisis)
        
        # KPIs
        col1, col2, col3, col4 = st.columns(4)
        
        with col1:
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Distritos Evaluados</div>
                <div class="kpi-value">{len(ranking_df)}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            muy_alto = len(ranking_df[ranking_df['Nivel_Riesgo'] == 'Muy Alto'])
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Riesgo Muy Alto</div>
                <div class="kpi-value" style="color: #ff6b6b;">{muy_alto}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            iria_promedio = ranking_df['IRIA'].mean()
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">IRIA Promedio</div>
                <div class="kpi-value">{iria_promedio:.2f}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col4:
            prob_promedio = ranking_df['Probabilidad'].mean()
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Probabilidad Promedio</div>
                <div class="kpi-value">{prob_promedio:.2%}</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Tabla y Gráfica lado a lado
        col_tabla, col_grafica = st.columns([1, 1.2])
        
        with col_tabla:
            st.markdown("### Tabla de Ranking")
            
            # Tabla formateada
            ranking_display = ranking_df.copy()
            ranking_display.index = ranking_display.index + 1
            ranking_display.index.name = 'Ranking'
            
            # Crear tabla HTML coloreada
            html_table = "<table style='width:100%; border-collapse: collapse;'>"
            html_table += "<tr style='background-color: #1a472a; color: white;'>"
            for col in ranking_display.columns:
                html_table += f"<th style='padding: 10px; text-align: left; border: 1px solid #ddd;'>{col}</th>"
            html_table += "</tr>"
            
            for idx, row in ranking_display.iterrows():
                color = nivel_iria_a_color(row['Nivel_Riesgo'])
                html_table += f"<tr style='background-color: {color}20;'>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: center;'><strong>{idx}</strong></td>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd;'>{row['Distrito']}</td>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: center;'><strong>{row['IRIA']:.4f}</strong></td>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: center;'><span style='background-color: {color}; color: white; padding: 4px 8px; border-radius: 4px;'>{row['Nivel_Riesgo']}</span></td>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: center;'>{row['Probabilidad']:.2%}</td>"
                html_table += f"<td style='padding: 8px; border: 1px solid #ddd; text-align: right;'>{row['Población']:,}</td>"
                html_table += "</tr>"
            
            html_table += "</table>"
            st.markdown(html_table, unsafe_allow_html=True)
        
        with col_grafica:
            st.markdown("### Análisis Visual")
            fig = crear_grafica_distribucion_iria(ranking_df)
            st.pyplot(fig, use_container_width=True)
        
        st.markdown("---")
        
        # Exportar datos
        col_exp1, col_exp2 = st.columns(2)
        
        with col_exp1:
            excel_data = exportar_a_excel(ranking_df, modelos)
            st.download_button(
                label="📥 Descargar Excel",
                data=excel_data,
                file_name=f"ranking_inseguridad_alimentaria_{año_analisis}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        
        with col_exp2:
            csv_data = ranking_df.to_csv(index=False).encode('utf-8')
            st.download_button(
                label="📥 Descargar CSV",
                data=csv_data,
                file_name=f"ranking_inseguridad_alimentaria_{año_analisis}.csv",
                mime="text/csv"
            )
    
    # ==================== PESTAÑA 2: PROYECCIONES FUTURAS ====================
    elif tab_tipo == "🔮 Proyecciones Futuras":
        st.markdown("## 🔮 Proyecciones de Inseguridad Alimentaria (2024-2035)")
        
        col1, col2 = st.columns(2)
        
        with col1:
            distrito_seleccionado = st.selectbox(
                "Selecciona un distrito:",
                sorted(df_normalizado['Distrito'].unique())
            )
        
        with col2:
            años_proyeccion = st.slider(
                "Rango de años a proyectar",
                min_value=2024,
                max_value=2035,
                value=(2024, 2035),
                step=1
            )
        
        # Proyectar
        proyecciones = proyectar_riesgo(
            df_normalizado,
            modelos,
            distrito_seleccionado,
            años_proyeccion[0],
            años_proyeccion[1]
        )
        
        if not proyecciones.empty:
            # KPIs de proyección
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                iria_actual = proyecciones[proyecciones['Año'] == años_proyeccion[0]]['IRIA'].values[0]
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">IRIA {años_proyeccion[0]}</div>
                    <div class="kpi-value">{iria_actual:.2f}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col2:
                iria_final = proyecciones[proyecciones['Año'] == años_proyeccion[1]]['IRIA'].values[0]
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">IRIA {años_proyeccion[1]}</div>
                    <div class="kpi-value">{iria_final:.2f}</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col3:
                cambio_iria = ((iria_final - iria_actual) / iria_actual * 100) if iria_actual > 0 else 0
                color_cambio = "#4caf50" if cambio_iria < 0 else "#f44336"
                st.markdown(f"""
                <div class="kpi-card" style="border-left-color: {color_cambio};">
                    <div class="kpi-label">Cambio IRIA</div>
                    <div class="kpi-value" style="color: {color_cambio};">{cambio_iria:+.1f}%</div>
                </div>
                """, unsafe_allow_html=True)
            
            with col4:
                nivel_inicial = proyecciones[proyecciones['Año'] == años_proyeccion[0]]['Nivel_Riesgo'].values[0]
                nivel_final = proyecciones[proyecciones['Año'] == años_proyeccion[1]]['Nivel_Riesgo'].values[0]
                cambio_nivel = "↑" if nivel_inicial == nivel_final else ("↑" if nivel_final > nivel_inicial else "↓")
                st.markdown(f"""
                <div class="kpi-card">
                    <div class="kpi-label">Cambio de Nivel {cambio_nivel}</div>
                    <div class="kpi-value" style="font-size: 1.2rem;">{nivel_inicial} → {nivel_final}</div>
                </div>
                """, unsafe_allow_html=True)
            
            st.markdown("---")
            
            # Tabla de proyecciones
            st.markdown("###  Tabla de Proyecciones")
            proyecciones_display = proyecciones[['Año', 'IRIA', 'Nivel_Riesgo', 'Probabilidad']].copy()
            proyecciones_display['IRIA'] = proyecciones_display['IRIA'].round(4)
            proyecciones_display['Probabilidad'] = proyecciones_display['Probabilidad'].round(4)
            
            st.dataframe(proyecciones_display, use_container_width=True, hide_index=True)
            
            # Gráficas
            col_g1, col_g2 = st.columns(2)
            
            with col_g1:
                fig, ax = plt.subplots(figsize=(10, 6))
                colors = [nivel_iria_a_color(nivel) for nivel in proyecciones['Nivel_Riesgo']]
                ax.plot(proyecciones['Año'], proyecciones['IRIA'], marker='o', linewidth=3, 
                       markersize=10, color='#2d8659', label='IRIA')
                ax.scatter(proyecciones['Año'], proyecciones['IRIA'], c=colors, s=200, 
                          edgecolors='black', linewidth=2, zorder=5)
                ax.fill_between(proyecciones['Año'], proyecciones['IRIA'], alpha=0.3, color='#2d8659')
                ax.set_xlabel('Año', fontweight='bold', fontsize=11)
                ax.set_ylabel('Índice IRIA', fontweight='bold', fontsize=11)
                ax.set_title(f'Proyección de IRIA - {distrito_seleccionado}', fontweight='bold', fontsize=12, color='#1a472a')
                ax.grid(True, alpha=0.3, linestyle='--')
                ax.set_ylim(0, 1)
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
            
            with col_g2:
                fig, ax = plt.subplots(figsize=(10, 6))
                ax.plot(proyecciones['Año'], proyecciones['Probabilidad'], marker='s', linewidth=3,
                       markersize=10, color='#ff9800', label='Probabilidad')
                ax.fill_between(proyecciones['Año'], proyecciones['Probabilidad'], alpha=0.3, color='#ff9800')
                ax.set_xlabel('Año', fontweight='bold', fontsize=11)
                ax.set_ylabel('Probabilidad', fontweight='bold', fontsize=11)
                ax.set_title(f'Proyección de Probabilidad - {distrito_seleccionado}', fontweight='bold', fontsize=12, color='#1a472a')
                ax.grid(True, alpha=0.3, linestyle='--')
                ax.set_ylim(0, 1)
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
        else:
            st.error("No se encontraron datos para el distrito seleccionado.")
    
    # ==================== PESTAÑA 3: ANÁLISIS DETALLADO ====================
    elif tab_tipo == " Análisis Detallado":
        st.markdown("##  Análisis Detallado por Distrito")
        
        distrito_analisis = st.selectbox(
            "Selecciona un distrito para análisis detallado:",
            sorted(df_normalizado['Distrito'].unique()),
            key="analisis_detallado"
        )
        
        distrito_data = df_normalizado[df_normalizado['Distrito'] == distrito_analisis].iloc[0]
        
        # Información general
        col1, col2, col3 = st.columns(3)
        
        with col1:
            iria = calcular_iria(
                distrito_data['Ingresos_Laborales'],
                distrito_data['Gasto_Alimentario'],
                distrito_data['Indice_Vulnerabilidad'],
                distrito_data['Indice_Privacion_Multidimensional']
            )
            nivel = nivel_iria_a_texto(iria)
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Índice IRIA</div>
                <div class="kpi-value">{iria:.2f}</div>
                <div style="margin-top: 10px; padding: 8px; background: {nivel_iria_a_color(nivel)}; border-radius: 5px; color: white; font-weight: bold;">{nivel}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col2:
            X_pred = np.array([[
                distrito_data['Ingresos_Laborales'],
                distrito_data['Gasto_Alimentario'],
                distrito_data['Indice_Vulnerabilidad'],
                distrito_data['Indice_Privacion_Multidimensional']
            ]])
            prob = modelos['regressor'].predict(X_pred)[0]
            prob = np.clip(prob, 0, 1)
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Probabilidad de Inseguridad</div>
                <div class="kpi-value">{prob:.1%}</div>
            </div>
            """, unsafe_allow_html=True)
        
        with col3:
            poblacion = int(distrito_data.get('Poblacion', 0))
            inseguros = int(poblacion * prob)
            st.markdown(f"""
            <div class="kpi-card">
                <div class="kpi-label">Población en Riesgo</div>
                <div class="kpi-value">{inseguros:,}</div>
                <div style="margin-top: 5px; font-size: 0.9rem; opacity: 0.9;">de {poblacion:,}</div>
            </div>
            """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Indicadores
        st.markdown("###  Indicadores del Distrito")
        
        indicadores = [
            ('Ingresos Laborales (Normalizado)', distrito_data['Ingresos_Laborales']),
            ('Gasto Alimentario (Normalizado)', distrito_data['Gasto_Alimentario']),
            ('Índice de Vulnerabilidad', distrito_data['Indice_Vulnerabilidad']),
            ('Índice de Privación Multidimensional', distrito_data['Indice_Privacion_Multidimensional'])
        ]
        
        for nombre, valor in indicadores:
            col_ind1, col_ind2 = st.columns([0.2, 0.8])
            
            with col_ind1:
                st.write(f"**{nombre}**")
            
            with col_ind2:
                # Barra de progreso coloreada
                color_ind = '#4caf50' if valor < 0.33 else ('#ff9800' if valor < 0.66 else '#f44336')
                st.markdown(f"""
                <div class="progress-bar">
                    <div class="progress-fill" style="width: {valor*100}%; background: linear-gradient(90deg, #2d8659, {color_ind});">
                        {valor:.2%}
                    </div>
                </div>
                """, unsafe_allow_html=True)
        
        st.markdown("---")
        
        # Proyección compacta
        st.markdown("### 🔮 Proyección a 2035")
        
        proyecciones_2035 = proyectar_riesgo(df_normalizado, modelos, distrito_analisis, 2024, 2035)
        
        if not proyecciones_2035.empty:
            col_p1, col_p2 = st.columns(2)
            
            with col_p1:
                fig, ax = plt.subplots(figsize=(8, 5))
                ax.plot(proyecciones_2035['Año'], proyecciones_2035['IRIA'], marker='o', 
                       linewidth=2.5, markersize=8, color='#2d8659')
                ax.fill_between(proyecciones_2035['Año'], proyecciones_2035['IRIA'], alpha=0.2, color='#2d8659')
                ax.set_xlabel('Año', fontweight='bold')
                ax.set_ylabel('IRIA', fontweight='bold')
                ax.set_title('Evolución IRIA 2024-2035', fontweight='bold', color='#1a472a')
                ax.grid(True, alpha=0.3, linestyle='--')
                ax.set_ylim(0, 1)
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
            
            with col_p2:
                fig, ax = plt.subplots(figsize=(8, 5))
                ax.plot(proyecciones_2035['Año'], proyecciones_2035['Probabilidad'], marker='s',
                       linewidth=2.5, markersize=8, color='#ff9800')
                ax.fill_between(proyecciones_2035['Año'], proyecciones_2035['Probabilidad'], alpha=0.2, color='#ff9800')
                ax.set_xlabel('Año', fontweight='bold')
                ax.set_ylabel('Probabilidad', fontweight='bold')
                ax.set_title('Evolución Probabilidad 2024-2035', fontweight='bold', color='#1a472a')
                ax.grid(True, alpha=0.3, linestyle='--')
                ax.set_ylim(0, 1)
                plt.tight_layout()
                st.pyplot(fig, use_container_width=True)
    
    # ==================== FOOTER ====================
    st.markdown("---")
    col_footer1, col_footer2, col_footer3 = st.columns(3)
    
    with col_footer1:
        st.caption(f" Última actualización: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
    
    with col_footer2:
        st.caption(f" Distritos analizados: {len(df_normalizado['Distrito'].unique())}")
    
    with col_footer3:
        st.caption(" Datos confidenciales - Uso exclusivo autorizado")

else:
    st.error(" No se pudo cargar el archivo de datos. Verifica que existe el archivo Excel.")
